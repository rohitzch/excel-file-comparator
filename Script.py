import openpyxl

def compare_excel_files(file1, file2):
    # Load workbooks
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Check if they have the same number of sheets
    if wb1.sheetnames != wb2.sheetnames:
        return False

    # Compare each sheet
    for sheet_name in wb1.sheetnames:
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        # Check if sheets have the same dimensions
        if sheet1.max_row != sheet2.max_row or sheet1.max_column != sheet2.max_column:
            return False

        # Compare cell values
        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                if sheet1.cell(row, col).value != sheet2.cell(row, col).value:
                    return False

    return True

# Usage
file1 = r'C:\Users\rohitzch\Desktop\500m UPDATEDhhh.xlsx'
file2 = r'C:\Users\rohitzch\Downloads\500m UPDATED.xlsx'

if compare_excel_files(file1, file2):
    print("The files are identical.")
else:
    print("The files are different.")
